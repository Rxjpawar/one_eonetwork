from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from zipfile import BadZipFile
import os

HEADERS = [
    "Personal Information",
    "Professional Experience",
    "Contact Information",
    "Family Information",
    "Summary"
]


def to_cell_value(value):
    """
    Safely convert any LLM output value to an Excel-safe string.
    Handles str, dict, list, None — whatever the LLM returns.
    """
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, dict):
        return "\n".join(
            f"{k}: {to_cell_value(v)}" for k, v in value.items()
        )
    if isinstance(value, list):
        return "\n\n".join(to_cell_value(item) for item in value)
    return str(value).strip()


def save_to_excel(rows, FILE_PATH):
    if os.path.exists(FILE_PATH):
        try:
            wb = load_workbook(FILE_PATH)
            ws = wb.active
        except BadZipFile:
            os.remove(FILE_PATH)
            wb = Workbook()
            ws = wb.active
            ws.append(HEADERS)
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(HEADERS)

    # Bold headers
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Insert newest rows at top (row 2, below header)
    for r in rows:
        ws.insert_rows(2)
        ws.cell(2, 1).value = to_cell_value(r.get("personal_information", ""))
        ws.cell(2, 2).value = to_cell_value(r.get("professional_experience", ""))
        ws.cell(2, 3).value = to_cell_value(r.get("contact_information", ""))
        ws.cell(2, 4).value = to_cell_value(r.get("family_information", ""))
        ws.cell(2, 5).value = to_cell_value(r.get("summary", ""))

    # Column widths
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 60

    # Wrap text + top-align all data rows
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(f"data/{FILE_PATH}")