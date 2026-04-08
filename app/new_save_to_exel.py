from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from zipfile import BadZipFile
import os

HEADERS = [
    "Personal Information",
    "Professional Experience",
    "Contact Information",
    "Family Information",
    "Summary"
]

def to_cell_value(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, dict):
        return "\n".join(f"{k}: {to_cell_value(v)}" for k, v in value.items())
    if isinstance(value, list):
        return "\n\n".join(to_cell_value(item) for item in value)
    return str(value).strip()

def save_to_excel(rows, filename):
    FILE_PATH = f"data/{filename}.xlsx"

    os.makedirs("data", exist_ok=True)

    if os.path.exists(FILE_PATH):
        try:
            wb = load_workbook(FILE_PATH)
            ws = wb.active
        except BadZipFile:
            os.remove(FILE_PATH)
            wb = Workbook()
            ws = wb.active
            ws.append(HEADERS)
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(HEADERS)

    # Style headers
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Append data
    for r in rows:
        ws.append([
            to_cell_value(r.get("personal_information", "")),
            to_cell_value(r.get("professional_experience", "")),
            to_cell_value(r.get("contact_information", "")),
            to_cell_value(r.get("family_information", "")),
            to_cell_value(r.get("summary", ""))
        ])

    # Formatting
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 60

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Safe save (prevents corruption)
    temp_path = FILE_PATH + ".tmp"
    wb.save(temp_path)
    os.replace(temp_path, FILE_PATH)

    print(f"Saved successfully: {FILE_PATH}")